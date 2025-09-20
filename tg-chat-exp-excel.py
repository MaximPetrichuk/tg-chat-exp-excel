#!/usr/bin/env python3
"""
=== Telegram Chat Exporter to Excel file v1.0 ===
Клиент командной строки Telegram для экспорта содеожимого чата в фвйл Excel за указанный период времени
Зависимости: telethon, openpyxl, python-dotenv
"""

from telethon import TelegramClient
from telethon.tl.functions.channels import GetForumTopicsRequest
import datetime
from openpyxl import Workbook
import os
from dotenv import load_dotenv

print("=== Telegram Chat Exporter to Excel file v1.0 ===")
print("Вывод сообщений по топикам и выгрузка в Excel\n")

# Загрузка переменных из .env
load_dotenv()

API_ID = os.getenv("API_ID")   
API_HASH = os.getenv("API_HASH") 
PHONE = os.getenv("PHONE")
CHAT_ID = int(os.getenv("CHAT_ID"))

client = TelegramClient("tg-chat-exp-excel", API_ID, API_HASH)

year = os.getenv("YEAR_DEFAULT")
month = os.getenv("MONTH_DEFAULT")

# --- запрос года и месяца ---
user_input = input(f"Введите год [YYYY] (или нажмите Enter для {year}): ")
year = int(user_input if user_input else year)
user_input = input(f"Введите месяц [1-12] (или нажмите Enter для {month}): ")
month = int(user_input if user_input else month)


# --- Делаем start_date и end_date timezone-aware (UTC) ---
start_date = datetime.datetime(year, month, 1, tzinfo=datetime.timezone.utc)
if month == 12:
    end_date = datetime.datetime(year + 1, 1, 1, tzinfo=datetime.timezone.utc) - datetime.timedelta(seconds=1)
else:
    end_date = datetime.datetime(year, month + 1, 1, tzinfo=datetime.timezone.utc) - datetime.timedelta(seconds=1)

# --- Функция выводит список всех доступных чатов с их ID ---
async def list_chats():
    """Показать список всех доступных чатов с их ID"""
    print("\n📋 Список доступных чатов:")
    async for dialog in client.iter_dialogs():
        print(f"{dialog.name}  -->  {dialog.id}")
    print("===================================")
    print("Скопируй ID нужного чата и пропиши в файле .env\n")

# --- Функция для определения ID топика ---
def get_topic_id(msg):
    if not msg.reply_to:
        return 0  # General
    if not getattr(msg.reply_to, "forum_topic", False):
        return 0  # General
    if getattr(msg.reply_to, "reply_to_top_id", None):
        return msg.reply_to.reply_to_top_id
    return getattr(msg.reply_to, "reply_to_msg_id", 0)

# --- Основная функция ---
async def main():
    await client.start(phone=PHONE)

    # --- Получаем чат ---
    try:
        chat = await client.get_entity(CHAT_ID)
    except Exception as e:
        print(f"\n❌ Не удалось найти чат: {CHAT_ID}")
        await list_chats()
        return
    
    print(f"\nСобираю сообщения из чата {chat.title}...")

    # --- Получаем топики ---
    try:
        topics_resp = await client(GetForumTopicsRequest(
            channel=chat,
            offset_date=None,
            offset_id=0,
            offset_topic=0,
            limit=200,
            q=None
        ))
        topics = {t.id: t.title for t in topics_resp.topics}
    except Exception as e:
        print("Не удалось получить список топиков:", e)
        topics = {}

    topics[0] = "General"

    # --- Собираем сообщения по топикам ---
    messages_by_topic = {tid: [] for tid in topics.keys()}
    total_messages = 0

    # --- Подготовка файла Excel ---
    wb = Workbook()
    # Удаляем стандартный лист, если он пуст
    default_sheet = wb.active
    if default_sheet and default_sheet.title == "Sheet":
        wb.remove(default_sheet)
    ws = wb.create_sheet(title="Все сообщения")
    ws.append(["Топик", "Автор", "Дата", "Сообщение"])

    async for msg in client.iter_messages(chat, offset_date=start_date, reverse=True):
        if not msg.message:
            continue

        # --- Фильтр по дате ---
        if not (start_date <= msg.date <= end_date):
            continue

        topic_id = get_topic_id(msg)
        messages_by_topic.setdefault(topic_id, []).append(msg)
        topic_name = topics.get(topic_id, f"Topic {topic_id}")
        sender = await msg.get_sender()
        author = f"{sender.first_name or ''} {sender.last_name or ''} (@{sender.username or ''})" if sender else "?"
        ws.append([topic_name, author, msg.date.strftime("%Y-%m-%d %H:%M"), msg.message])
        total_messages += 1

    # --- Вывод на экран и запись в файл Excel ---
    if total_messages > 0:
        # Создаём лист Excel для списка топиков
        ws_list = wb.create_sheet(title="Список")
        ws_list.append(["Список топиков"])
        ws_list.append(["ID", "Топик", "Название листа", "Кол-во сообщений"])

        for tid, msgs in messages_by_topic.items():
            if not msgs:
                continue
            topic_name = topics.get(tid, f"Topic {tid}")
            print(f"\n===== Топик: {topic_name} ({len(msgs)} сообщений) =====\n")
            ws_list.append([tid, topic_name, topic_name[:31], len(msgs)])
            # Создаём лист Excel для топика
            ws = wb.create_sheet(title=topic_name[:31])  # имя листа ограничено 31 символом
            ws.append([topic_name])
            ws.append(["Автор", "Дата", "Сообщение"])

            for msg in msgs:
                sender = await msg.get_sender()
                author = f"{sender.first_name or ''} {sender.last_name or ''} (@{sender.username or ''})" if sender else "?"
                print(f"[{msg.date:%Y-%m-%d %H:%M}] {author}: {msg.message}")
                ws.append([author, msg.date.strftime("%Y-%m-%d %H:%M"), msg.message])

        filename = f"tg_messages_{chat.title}_{year}_{month:02d}.xlsx"
        wb.save(filename)
        print(f"\nОбщее количество найденных сообщений за {year}-{month:02d}: {total_messages}")
        print(f"\nСообщения сохранены в файл: {filename}")
    else:
        print("Файл Excel не создан, так как сообщений за этот месяц нет.")

    wb.close()


with client:
    client.loop.run_until_complete(main())
