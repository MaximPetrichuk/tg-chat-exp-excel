#!/usr/bin/env python3
"""
Core logic for Telegram Chat Exporter
"""

import asyncio
import datetime
import os
import re
from openpyxl import Workbook
from telethon import TelegramClient
from telethon.tl.functions.channels import GetForumTopicsRequest
from dotenv import load_dotenv, set_key, dotenv_values

# --- Версия программы ---
PROGRAM_NAME = "Telegram Chat Exporter"
PROGRAM_VERSION = "v1.1"

SESSION_NAME = "tg_session"

# --- Работа с .env ---
ENV_FILE = ".env"


def load_env_vars():
    """Загружает переменные окружения из .env и возвращает словарь."""
    load_dotenv(ENV_FILE)
    cfg = dotenv_values(ENV_FILE)
    return {
        "API_ID": os.getenv("API_ID") or cfg.get("API_ID", ""),
        "API_HASH": os.getenv("API_HASH") or cfg.get("API_HASH", ""),
        "PHONE": os.getenv("PHONE") or cfg.get("PHONE", ""),
        "YEAR_DEFAULT": os.getenv("YEAR_DEFAULT") or cfg.get("YEAR_DEFAULT", ""),
        "MONTH_DEFAULT": os.getenv("MONTH_DEFAULT") or cfg.get("MONTH_DEFAULT", ""),
    }


def save_env_vars(api_id, api_hash, phone, year=None, month=None):
    """Сохраняет переменные окружения в .env."""
    def _set(k, v):
        if v:
            set_key(ENV_FILE, k, str(v))

    _set("API_ID", api_id)
    _set("API_HASH", api_hash)
    _set("PHONE", phone)
    if year:
        _set("YEAR_DEFAULT", year)
    if month:
        _set("MONTH_DEFAULT", month)


def check_env_vars(required=("API_ID", "API_HASH", "PHONE")):
    """
    Проверяет наличие обязательных параметров в окружении (.env).
    Возвращает (ok: bool, missing: list).
    """
    env = load_env_vars()
    missing = [key for key in required if not env.get(key)]
    return (len(missing) == 0, missing)


# --- Общие утилиты ---
def _safe_filename(s: str) -> str:
    if not s:
        return "chat"
    return re.sub(r'[\\/*?:"<>|]', "_", s)


def create_telegram_client(api_id, api_hash, session_name, log_callback=None):
    """
    Создает экземпляр TelegramClient с новым event loop (для потоков).
    """
    if log_callback is None:
        log_callback = lambda s: None

    try:
        api_id = int(api_id)
    except Exception:
        raise ValueError("API_ID должен быть числом")

    client = TelegramClient(session_name, api_id, api_hash)

    # Устанавливаем loop для клиента вручную
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    client._loop = loop

    return client


# ----------------- LIST CHATS -----------------
async def _list_chats_async(client, phone, log_callback):
    chats = []
    try:
        log_callback("Подключение к Telegram для получения списка чатов...")
        await client.start(phone=phone)
        log_callback("Подключение установлено. Получаю список чатов...")
        async for dialog in client.iter_dialogs():
            name = getattr(dialog, "name", str(dialog.entity))
            cid = dialog.id
            chats.append((name, cid))
        log_callback(f"Найдено {len(chats)} чатов.")
        return chats
    finally:
        try:
            await client.disconnect()
            log_callback("Отключился от Telegram (list_chats).")
        except Exception:
            pass


def list_chats(api_id, api_hash, session_name, phone, log_callback=None):
    """Синхронная функция. Возвращает список [(name, id), ...]."""
    if log_callback is None:
        log_callback = lambda s: print(s)

    client = create_telegram_client(api_id, api_hash, session_name, log_callback)
    loop = client._loop
    try:
        return loop.run_until_complete(_list_chats_async(client, phone, log_callback))
    except Exception as e:
        log_callback(f"Ошибка при получении списка чатов: {e}")
        return []
    finally:
        try:
            loop.close()
        except Exception:
            pass


# ----------------- EXPORT MESSAGES -----------------
async def _export_messages_async(client, phone, chat_id, year, month, log_callback):
    try:
        log_callback("Подключение к Telegram...")
        await client.start(phone=phone)
        log_callback("Подключение установлено.")

        log_callback(f"Получение информации о чате (ID={chat_id})...")
        try:
            chat = await client.get_entity(int(chat_id))
        except Exception as e:
            return {"success": False, "message": f"Не удалось найти чат {chat_id}: {e}"}

        chat_title = getattr(chat, "title", str(chat))
        log_callback(f"Чат: {chat_title}")

        # даты
        try:
            start_date = datetime.datetime(year, month, 1, tzinfo=datetime.timezone.utc)
            if month == 12:
                end_date = datetime.datetime(year + 1, 1, 1, tzinfo=datetime.timezone.utc) - datetime.timedelta(seconds=1)
            else:
                end_date = datetime.datetime(year, month + 1, 1, tzinfo=datetime.timezone.utc) - datetime.timedelta(seconds=1)
        except Exception as e:
            return {"success": False, "message": f"Неверный год/месяц: {e}"}

        # топики
        log_callback("Получение списка топиков (если есть)...")
        try:
            topics_resp = await client(GetForumTopicsRequest(
                channel=chat,
                offset_date=None,
                offset_id=0,
                offset_topic=0,
                limit=200,
                q=None
            ))
            topics = {t.id: t.title for t in topics_resp.topics}
            log_callback(f"Топиков получено: {len(topics)}")
        except Exception as e:
            log_callback(f"Не удалось получить топики: {e}")
            topics = {}
        topics[0] = "General"

        messages_by_topic = {tid: [] for tid in topics.keys()}
        total_messages = 0
        log_counter = 0

        log_callback("Сбор сообщений...")
        async for msg in client.iter_messages(chat, offset_date=start_date, reverse=True):
            if not msg.message:
                continue
            if not (start_date <= msg.date <= end_date):
                continue
            # определяем topic_id
            if not msg.reply_to:
                topic_id = 0
            elif not getattr(msg.reply_to, "forum_topic", False):
                topic_id = 0
            elif getattr(msg.reply_to, "reply_to_top_id", None):
                topic_id = msg.reply_to.reply_to_top_id
            else:
                topic_id = getattr(msg.reply_to, "reply_to_msg_id", 0)

            sender = await msg.get_sender()
            author = f"{(sender.first_name or '')} {(sender.last_name or '')}".strip()
            if sender and getattr(sender, "username", None):
                author += f" (@{sender.username})"
            messages_by_topic.setdefault(topic_id, []).append((author or "?", msg.date, msg.message))
            total_messages += 1
            log_counter += 1
            if log_counter % 100 == 0:
                log_callback(f"Прочитано {total_messages} сообщений...")

        if total_messages == 0:
            log_callback("Сообщений за выбранный период не найдено.")
            return {"success": False, "message": "Сообщений за выбранный период нет."}

        log_callback(f"Найдено {total_messages} сообщений. Подготовка Excel-файла...")

        wb = Workbook()
        default_sheet = wb.active
        if default_sheet and default_sheet.title == "Sheet":
            wb.remove(default_sheet)

        ws_all = wb.create_sheet(title="Все сообщения")
        ws_all.append(["Топик", "Автор", "Дата", "Сообщение"])

        for tid, msgs in messages_by_topic.items():
            tname = topics.get(tid, f"Topic {tid}")
            for author, date_obj, text in msgs:
                ws_all.append([tname, author, date_obj.strftime("%Y-%m-%d %H:%M"), text])

        ws_list = wb.create_sheet(title="Список")
        ws_list.append(["ID", "Топик", "Название листа", "Кол-во сообщений"])
        for tid, msgs in messages_by_topic.items():
            if not msgs:
                continue
            tname = topics.get(tid, f"Topic {tid}")
            ws_list.append([tid, tname, tname[:31], len(msgs)])
            ws_topic = wb.create_sheet(title=tname[:31])
            ws_topic.append([tname])
            ws_topic.append(["Автор", "Дата", "Сообщение"])
            for author, date_obj, text in msgs:
                ws_topic.append([author, date_obj.strftime("%Y-%m-%d %H:%M"), text])

        safe_title = _safe_filename(chat_title)
        filename = f"tg_messages_{safe_title}_{year}_{month:02d}.xlsx"
        log_callback("Запись в файл Excel...")
        wb.save(filename)
        wb.close()
        abs_path = os.path.abspath(filename)
        log_callback(f"Файл сохранён: {abs_path}")

        return {"success": True, "filename": abs_path, "count": total_messages}

    finally:
        try:
            await client.disconnect()
            log_callback("Отключение от Telegram.")
        except Exception:
            pass


def export_messages(api_id, api_hash, session_name, phone, chat_id, year, month, log_callback=None):
    """Синхронная обёртка для экспорта сообщений."""
    if log_callback is None:
        log_callback = lambda s: print(s)

    client = create_telegram_client(api_id, api_hash, session_name, log_callback)
    loop = client._loop
    try:
        return loop.run_until_complete(
            _export_messages_async(client, phone, chat_id, year, month, log_callback)
        )
    except Exception as e:
        log_callback(f"Исключение при экспорте: {e}")
        return {"success": False, "message": f"Исключение: {e}"}
    finally:
        try:
            loop.close()
        except Exception:
            pass
